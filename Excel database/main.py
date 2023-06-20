import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

if not os.path.exists("excle.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="Name")
    sheet.cell(row=1, column=2, value="Course")
    sheet.cell(row=1, column=3, value="Semester")
    sheet.cell(row=1, column=4, value="Form Number")
    sheet.cell(row=1, column=5, value="Contact Number")
    sheet.cell(row=1, column=6, value="Email")
    sheet.cell(row=1, column=7, value="Address")
    workbook.save("excle.xlsx")


workbook = load_workbook("excle.xlsx")
sheet = workbook.active


def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email"
    sheet.cell(row=1, column=7).value = "Address"

def focus1(event):
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()

def clear():
    name_field.delete(0, tk.END)
    course_field.delete(0, tk.END)
    sem_field.delete(0, tk.END)
    form_no_field.delete(0, tk.END)
    contact_no_field.delete(0, tk.END)
    email_id_field.delete(0, tk.END)
    address_field.delete(0, tk.END)

def insert():
    if(name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
        messagebox.showerror("Empty input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row=current_row+1, column=1).value = name_field.get()
        sheet.cell(row=current_row+1, column=2).value = course_field.get()
        sheet.cell(row=current_row+1, column=3).value = sem_field.get()
        sheet.cell(row=current_row+1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row+1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row+1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row+1, column=7).value = address_field.get()
        workbook.save("excle.xlsx")
        name_field.focus_set()
        clear()

root = tk.Tk()
root.configure(background="light green")
root.title("Registration Form")
root.geometry("500x300")

excel()

heading = tk.Label(root, text="Form", bg="light green")
name = tk.Label(root, text="Name", bg="light green")
course = tk.Label(root, text="Course", bg="light green")
sem = tk.Label(root, text="Semester", bg="light green")
form_no = tk.Label(root, text="Form Number", bg="light green")
contact_no = tk.Label(root, text="contact Number", bg="light green")
email_id = tk.Label(root, text="Email", bg="light green")
address = tk.Label(root, text="Address", bg="light green")

name_field = tk.Entry(root)
course_field = tk.Entry(root)
sem_field = tk.Entry(root)
form_no_field = tk.Entry(root)
contact_no_field = tk.Entry(root)
email_id_field = tk.Entry(root)
address_field = tk.Entry(root)

name_field.bind("<Return>", focus1)
course_field.bind("<Return>", focus2)
sem_field.bind("<Return>", focus3)
form_no_field.bind("<Return>", focus4)
contact_no_field.bind("<Return>", focus5)
email_id_field.bind("<Return>", focus6)

heading.grid(row=0, column=1)
name.grid(row=1, column=0)
course.grid(row=2, column=0)
sem.grid(row=3, column=0)
form_no.grid(row=4, column=0)
contact_no.grid(row=5, column=0)
email_id.grid(row=6, column=0)
address.grid(row=7, column=0)

name_field.grid(row=1, column=1, ipadx=100)
course_field.grid(row=2, column=1, ipadx=100)
sem_field.grid(row=3, column=1, ipadx=100)
form_no_field.grid(row=4, column=1, ipadx=100)
contact_no_field.grid(row=5, column=1, ipadx=100)
email_id_field.grid(row=6, column=1, ipadx=100)
address_field.grid(row=7, column=1, ipadx=100)


submit = tk.Button(root, text="Submit", fg="black", bg="red", command=insert)
submit.grid(row=8, column=1)

root.mainloop()

